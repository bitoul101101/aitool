"""
CSV and Excel reporter.
"""

import csv
import os
from pathlib import Path
from typing import List, Dict, Any


EXPORT_COLUMNS = [
    "repo", "ai_category", "provider_or_lib",
    "capability", "policy_status", "risk", "severity",
    "file", "line", "snippet", "owner", "last_seen", "remediation",
]


class CSVReporter:

    def __init__(self, output_dir: str, scan_id: str):
        self.output_dir = Path(output_dir)
        self.scan_id = scan_id

    def write_csv(self, findings: List[Dict[str, Any]]) -> str:
        path = self.output_dir / f"ai_scan_{self.scan_id}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for finding in findings:
                # Sanitize snippet for CSV
                row = dict(finding)
                row["snippet"] = _sanitize(row.get("snippet", ""))
                writer.writerow(row)
        return str(path)

    def write_excel(self, findings: List[Dict[str, Any]]) -> str:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback: write CSV with .xlsx extension note
            path = self.output_dir / f"ai_scan_{self.scan_id}_note.txt"
            path.write_text("Install openpyxl for Excel output: pip install openpyxl")
            return str(path)

        wb = openpyxl.Workbook()

        # ── Main findings sheet ──────────────────────────────────
        ws = wb.active
        ws.title = "AI Findings"

        SEV_COLORS = {1: "C00000", 2: "FF0000", 3: "FFC000", 4: "FFFF00"}
        POLICY_COLORS = {
            "CRITICAL": "C00000", "BANNED": "C00000",
            "RESTRICTED": "FF0000", "REVIEW": "FFC000",
            "ALLOWED": "70AD47", "APPROVED": "70AD47",
        }

        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        ws.row_dimensions[1].height = 30

        # Data rows
        for row_idx, finding in enumerate(findings, 2):
            sev = finding.get("severity", 4)
            policy = finding.get("policy_status", "REVIEW")

            for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
                val = finding.get(col_name, "")
                if col_name == "snippet":
                    val = _sanitize(str(val))[:500]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("snippet", "remediation")))
                cell.font = Font(size=9)

                # Color severity column
                if col_name == "severity":
                    color = SEV_COLORS.get(sev, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(bold=True, size=9,
                                     color="FFFFFF" if sev <= 2 else "000000")

                # Color policy status column
                if col_name == "policy_status":
                    color = POLICY_COLORS.get(policy, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(bold=True, size=9,
                                     color="FFFFFF" if policy in ("CRITICAL","BANNED","RESTRICTED") else "000000")

        # Column widths
        col_widths = {
            "repo": 20, "ai_category": 18,
            "provider_or_lib": 22, "capability": 20, "policy_status": 14,
            "risk": 10, "severity": 10, "file": 40, "line": 8,
            "snippet": 50, "owner": 18, "last_seen": 16, "remediation": 60,
        }
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ── Summary sheet ────────────────────────────────────────
        from collections import Counter
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15

        def _sh(row, a, b, bold=False):
            c1 = ws2.cell(row=row, column=1, value=a)
            c2 = ws2.cell(row=row, column=2, value=b)
            if bold:
                c1.font = Font(bold=True)
                c2.font = Font(bold=True)

        r = 1
        _sh(r, "AI Usage Scan Summary", "", bold=True); r += 1
        _sh(r, "Scan ID", self.scan_id); r += 1
        _sh(r, "Total Findings", len(findings)); r += 1
        r += 1

        _sh(r, "By Severity", "", bold=True); r += 1
        sev_counter = Counter(f["severity"] for f in findings)
        for sev in sorted(sev_counter):
            labels = {1:"Critical",2:"High",3:"Medium",4:"Low"}
            _sh(r, f"Sev-{sev} ({labels[sev]})", sev_counter[sev]); r += 1
        r += 1

        _sh(r, "By Policy Status", "", bold=True); r += 1
        ps_counter = Counter(f.get("policy_status","REVIEW") for f in findings)
        for k, v in ps_counter.most_common():
            _sh(r, k, v); r += 1
        r += 1

        _sh(r, "By Category", "", bold=True); r += 1
        cat_counter = Counter(f["ai_category"] for f in findings)
        for k, v in cat_counter.most_common():
            _sh(r, k, v); r += 1
        r += 1

        path = self.output_dir / f"ai_scan_{self.scan_id}.xlsx"
        wb.save(str(path))
        return str(path)


def _sanitize(text: str) -> str:
    """Remove control characters that break CSV/Excel."""
    return "".join(c if c.isprintable() or c in ("\n", "\t") else " " for c in text)
